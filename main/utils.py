# utils.py
import openpyxl

def export_items_to_excel(items, response):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Add headers
    headers = ['Name', 'Description', 'Consumable', 'Quantity', 'Condition', 'Assigned Office', 'Store Entered Date', 'Office Assigned Date', 'Store Returned Date', 'Damaged Date', 'Maintained Date']
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Add data
    for row_num, item in enumerate(items, 2):
        worksheet.cell(row=row_num, column=1, value=item.name)
        worksheet.cell(row=row_num, column=2, value=item.description)
        # Add other fields similarly

    workbook.save(response)

from .group_mappings import GROUP_NAME_MAPPING

def get_user_group_full_name(user):
    user_group_name = user.groups.first().name if user.groups.exists() else None
    
    return GROUP_NAME_MAPPING.get(user_group_name, user_group_name)
